#!/usr/bin/env python3
"""
Teste une liste de domaines pour determiner s'ils hebergent un site web
operationnel ou si le domaine sert uniquement a la messagerie (MX only).

Dependances: pip install dnspython requests openpyxl

Usage:
    python3 test_domaines.py --excel BMR_prospection.xlsx --col Domaine --out rapport_domaines.csv
    python3 test_domaines.py --domaines domaine1.com domaine2.ca
"""

import argparse
import csv
import sys
import time

import dns.resolver
import requests

USER_AGENT = "Mozilla/5.0 (compatible; DomainAudit/1.0)"
TIMEOUT = 10

# Signatures textuelles typiques des pages de parking / domaines a vendre
PARKING_SIGNALS = [
    "domain is for sale", "buy this domain", "domain parking",
    "this domain may be for sale", "sedo", "godaddy.com/domainfind",
    "the domain has expired", "namecheap parking", "future home of something",
    "parked free", "this web page is parked",
]


def check_dns(domain):
    """Retourne (has_a, has_mx) en interrogeant les enregistrements A/AAAA et MX."""
    has_a = False
    has_mx = False
    for rtype in ("A", "AAAA"):
        try:
            dns.resolver.resolve(domain, rtype, lifetime=TIMEOUT)
            has_a = True
            break
        except Exception:
            continue
    try:
        dns.resolver.resolve(domain, "MX", lifetime=TIMEOUT)
        has_mx = True
    except Exception:
        pass
    return has_a, has_mx


def check_http(domain):
    """Essaie https:// puis http://, avec et sans www.
    Retourne (status_code, is_real_site, error)."""
    candidates = [domain, f"www.{domain}"] if not domain.startswith("www.") else [domain]
    for host in candidates:
        for scheme in ("https", "http"):
            url = f"{scheme}://{host}"
            try:
                r = requests.get(
                    url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT},
                    allow_redirects=True,
                )
                text_lower = r.text.lower()
                is_parked = any(sig in text_lower for sig in PARKING_SIGNALS)
                is_too_thin = len(r.text.strip()) < 200
                is_real_site = r.status_code < 400 and not is_parked and not is_too_thin
                return r.status_code, is_real_site, None
            except requests.exceptions.SSLError:
                continue  # essaie http:// ou l'hote suivant
            except Exception as e:
                last_err = str(e)
                continue
    return None, False, locals().get("last_err", "aucune reponse")


def classify(has_a, has_mx, http_status, is_real_site):
    if is_real_site:
        return "Site web actif"
    if has_a and http_status is not None:
        return "Domaine actif, site inaccessible ou parque"
    if has_mx and not has_a:
        return "Emails seulement (MX sans site)"
    if has_mx and has_a:
        return "DNS actif (mail+A) mais site non confirme"
    if not has_a and not has_mx:
        return "Domaine introuvable ou inactif"
    return "Indetermine"


def test_domain(domain):
    domain = domain.strip().lower().replace("https://", "").replace("http://", "").rstrip("/")
    has_a, has_mx = check_dns(domain)
    http_status, is_real_site, err = check_http(domain)
    verdict = classify(has_a, has_mx, http_status, is_real_site)
    return {
        "domaine": domain,
        "a_record": has_a,
        "mx_record": has_mx,
        "http_status": http_status,
        "site_reel_detecte": is_real_site,
        "erreur_http": err,
        "verdict": verdict,
    }


def load_domains_from_excel(path, column):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = headers.index(column) + 1
    domains = []
    for i in range(2, ws.max_row + 1):
        v = ws.cell(i, idx).value
        if v:
            domains.append(v.strip())
    return domains


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--excel", help="Fichier Excel source")
    ap.add_argument("--col", default="Domaine", help="Nom de la colonne contenant les domaines")
    ap.add_argument("--domaines", nargs="*", help="Liste de domaines a tester directement")
    ap.add_argument("--out", default="rapport_domaines.csv", help="Fichier CSV de sortie")
    ap.add_argument("--delay", type=float, default=0.5, help="Delai entre requetes (secondes)")
    args = ap.parse_args()

    if args.excel:
        domains = load_domains_from_excel(args.excel, args.col)
    elif args.domaines:
        domains = args.domaines
    else:
        ap.error("Fournir --excel ou --domaines")

    unique_domains = sorted(set(d.lower() for d in domains))
    print(f"{len(unique_domains)} domaines uniques a tester")

    results = []
    for i, d in enumerate(unique_domains, 1):
        print(f"[{i}/{len(unique_domains)}] {d} ...", end=" ", flush=True)
        r = test_domain(d)
        print(r["verdict"])
        results.append(r)
        time.sleep(args.delay)

    with open(args.out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)

    print(f"\nRapport ecrit dans {args.out}")
    from collections import Counter
    counts = Counter(r["verdict"] for r in results)
    for verdict, n in counts.most_common():
        print(f"  {verdict}: {n}")


if __name__ == "__main__":
    main()
