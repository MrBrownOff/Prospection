"""
Enrichissement Google Maps — Liste de prospection Interbois
===========================================================
Script pour Claude Code.

Objectif : Enrichir la liste de 339 prospects avec le nombre d'avis Google
et la note, comme proxy de taille des établissements.

Utilisation :
    export RAPIDAPI_KEY=a61db8c9bcmshe9d2a1d0b97d50cp13f8dfjsn4feca1b3b610
    python enrichir_prospects_google_maps.py

Ou avec la clé en argument :
    python enrichir_prospects_google_maps.py --key ta_cle_ici

Fonctionnalités :
    - Checkpoint automatique : reprend là où ça s'est arrêté
    - Rate limiting : 1 req/seconde par défaut (ajustable)
    - Matching par coordonnées GPS (+ fallback par nom)
    - Génère un Excel enrichi final

API utilisée : google-maps-extractor2.p.rapidapi.com
"""

import os
import sys
import time
import json
import argparse
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── CONFIG ───────────────────────────────────────────────────────────────────
EXCEL_INPUT   = "Prospection_Interbois_Bannières_QC.xlsx"
CHECKPOINT    = "checkpoint_gmaps.json"
OUTPUT_EXCEL  = "Prospection_Interbois_Enrichi.xlsx"
OUTPUT_CSV    = "prospects_enrichis.csv"

API_HOST      = "google-maps-extractor2.p.rapidapi.com"
API_URL       = f"https://{API_HOST}/locate_and_search"
RATE_LIMIT    = 1.0   # secondes entre chaque requête (ajuster selon ta limite)
MAX_RESULTS   = 5     # nombre de résultats Google Maps à analyser par prospect
RADIUS_M      = 500   # rayon de recherche en mètres autour des coordonnées

# ── ARGUMENTS ────────────────────────────────────────────────────────────────
def parse_args():
    parser = argparse.ArgumentParser(description='Enrichissement Google Maps — Interbois')
    parser.add_argument('--key', help='Clé RapidAPI (sinon via env RAPIDAPI_KEY)')
    parser.add_argument('--rate', type=float, default=RATE_LIMIT,
                        help=f'Délai entre requêtes en secondes (défaut: {RATE_LIMIT})')
    parser.add_argument('--limit', type=int, default=0,
                        help='Limiter à N prospects (0 = tous) — utile pour tester')
    parser.add_argument('--reset', action='store_true',
                        help='Ignorer le checkpoint et recommencer depuis le début')
    parser.add_argument('--debug', action='store_true',
                        help='Afficher la réponse brute de l\'API pour diagnostic')
    return parser.parse_args()

# ── CHECKPOINT ───────────────────────────────────────────────────────────────
def load_checkpoint():
    if Path(CHECKPOINT).exists():
        with open(CHECKPOINT, 'r') as f:
            return json.load(f)
    return {}

def save_checkpoint(data):
    with open(CHECKPOINT, 'w') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── API CALL ─────────────────────────────────────────────────────────────────
def build_query(row):
    """Construit la requête de recherche Google Maps."""
    nom = str(row.get('Nom', '')).strip()
    ville = str(row.get('Ville', '')).strip()
    banniere = str(row.get('Bannière', '')).strip()

    # Nettoyer les préfixes de bannière dans le nom pour éviter la redondance
    nom_clean = nom
    for prefix in ['BMR ', 'Home Hardware ', 'Castle ', 'Timber Mart ']:
        if nom_clean.startswith(prefix):
            nom_clean = nom_clean[len(prefix):]

    # Requête : nom complet + ville + province
    query = f"{nom} {ville} Quebec Canada"
    return query

def search_gmaps(query, lat, lon, api_key, radius=RADIUS_M):
    """Appelle l'API Google Maps Extractor (locate_and_search, GET)."""
    headers = {
        "x-rapidapi-host": API_HOST,
        "x-rapidapi-key": api_key,
        "Content-Type": "application/json"
    }

    params = {
        "query":    query,
        "offset":   0,
        "limit":    MAX_RESULTS,
        "country":  "ca",
        "language": "fr",
    }

    try:
        response = requests.get(API_URL, params=params, headers=headers, timeout=15)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.HTTPError as e:
        if response.status_code == 429:
            print(f"  ⚠️  Rate limit atteint. Pause 60s...")
            time.sleep(60)
            return None
        elif response.status_code == 403:
            print(f"  ❌ Clé API invalide ou quota épuisé.")
            return None
        else:
            print(f"  ⚠️  HTTP {response.status_code}: {e}")
            return None
    except requests.exceptions.Timeout:
        print(f"  ⚠️  Timeout")
        return None
    except Exception as e:
        print(f"  ⚠️  Erreur: {e}")
        return None

def extract_best_match(results, nom_prospect, lat_prospect, lon_prospect):
    """
    Sélectionne le meilleur résultat parmi les résultats Google Maps.
    Stratégie : proximité GPS d'abord, puis similarité de nom.
    """
    if not results:
        return None

    # Normaliser les données selon la structure de l'API
    items = results if isinstance(results, list) else results.get('data', results.get('results', []))
    if not items:
        return None

    best = None
    best_score = -1

    for item in items[:MAX_RESULTS]:
        score = 0

        # Score par proximité GPS
        item_lat = item.get('latitude') or item.get('lat')
        item_lon = item.get('longitude') or item.get('lng')
        if item_lat and item_lon and lat_prospect and lon_prospect:
            try:
                dist = ((float(item_lat) - float(lat_prospect))**2 +
                        (float(item_lon) - float(lon_prospect))**2) ** 0.5
                if dist < 0.01:   # ~1km
                    score += 10
                elif dist < 0.05: # ~5km
                    score += 5
            except (ValueError, TypeError):
                pass

        # Score par similarité de nom (mots communs)
        item_name = str(item.get('name', item.get('business_name', ''))).upper()
        nom_mots = set(nom_prospect.upper().split())
        item_mots = set(item_name.split())
        mots_communs = nom_mots & item_mots
        score += len(mots_communs) * 2

        if score > best_score:
            best_score = score
            best = item

    return best if best_score > 0 else (items[0] if items else None)

def extract_metrics(item):
    """Extrait les métriques utiles d'un résultat Google Maps."""
    if not item:
        return {'gmaps_avis': None, 'gmaps_note': None,
                'gmaps_nom': None, 'gmaps_adresse': None,
                'gmaps_place_id': None, 'gmaps_url': None}

    return {
        'gmaps_avis':     item.get('reviews_count'),
        'gmaps_note':     item.get('rating'),
        'gmaps_nom':      item.get('name'),
        'gmaps_adresse':  item.get('full_address') or item.get('address'),
        'gmaps_place_id': item.get('place_id'),
        'gmaps_url':      item.get('website_url'),
    }

# ── SCORE TAILLE ─────────────────────────────────────────────────────────────
def score_taille(nb_avis):
    """
    Convertit le nombre d'avis en score de taille (1-5).
    Seuils calibrés pour les quincailleries/centres de rénovation QC.
    """
    if nb_avis is None:
        return None
    nb_avis = int(nb_avis) if str(nb_avis).isdigit() else 0
    if nb_avis >= 500:  return 5  # Grand format — volume fort
    if nb_avis >= 200:  return 4  # Format moyen-grand
    if nb_avis >= 100:  return 3  # Format moyen
    if nb_avis >= 50:   return 2  # Petit-moyen
    return 1                      # Petit — attention à la pertinence

def label_taille(score):
    labels = {5: '🟢 Grand (500+ avis)', 4: '🔵 Moyen-grand (200-499)',
              3: '🟡 Moyen (100-199)', 2: '🟠 Petit-moyen (50-99)',
              1: '🔴 Petit (<50 avis)', None: '⚪ Non enrichi'}
    return labels.get(score, '⚪ Non enrichi')

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    args = parse_args()

    # Clé API
    api_key = args.key or os.environ.get('RAPIDAPI_KEY')
    if not api_key:
        print("❌ Clé RapidAPI manquante. Utilise --key ou export RAPIDAPI_KEY=...")
        sys.exit(1)

    # Charger les prospects depuis l'Excel
    print(f"\n📂 Chargement de {EXCEL_INPUT}...")
    try:
        df = pd.read_excel(EXCEL_INPUT, sheet_name='🎯 Prospects', header=1)
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        print(f"❌ Impossible de lire {EXCEL_INPUT}: {e}")
        sys.exit(1)

    # Renommer colonnes pour correspondre au CSV interne
    col_map = {
        'Bannière': 'Bannière', 'Groupe / Propriétaire': 'Groupe',
        'Nom magasin': 'Nom', 'Ville': 'Ville', 'Adresse': 'Adresse',
        'Région': 'Région', 'Score territoire\n(1-5)': 'Score_territoire',
        'Format': 'Format', 'Téléphone': 'Téléphone',
        'URL magasin': 'URL_magasin', 'URL Google Maps': 'URL_GoogleMaps',
    }
    df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)

    # Ajouter colonnes GPS si absentes (chercher dans les colonnes disponibles)
    if 'Latitude' not in df.columns:
        df['Latitude'] = None
    if 'Longitude' not in df.columns:
        df['Longitude'] = None

    # Limiter si demandé (mode test)
    if args.limit > 0:
        df = df.head(args.limit)
        print(f"⚠️  Mode test : {args.limit} prospects seulement")

    total = len(df)
    print(f"✅ {total} prospects chargés\n")

    # Checkpoint
    checkpoint = {} if args.reset else load_checkpoint()
    deja_faits = len(checkpoint)
    if deja_faits > 0:
        print(f"📍 Checkpoint trouvé : {deja_faits} prospects déjà enrichis, reprise...\n")

    # Colonnes résultat
    cols_result = ['gmaps_avis', 'gmaps_note', 'score_taille', 'label_taille',
                   'gmaps_nom', 'gmaps_adresse', 'gmaps_place_id', 'gmaps_url']
    for col in cols_result:
        if col not in df.columns:
            df[col] = None

    # ── BOUCLE PRINCIPALE ────────────────────────────────────────────────────
    nouveaux = 0
    erreurs = 0

    for idx, row in df.iterrows():
        nom = str(row.get('Nom', '')).strip()
        key_cp = f"{row.get('Bannière','')}|{nom}|{row.get('Ville','')}"

        # Déjà dans le checkpoint ?
        if key_cp in checkpoint and not args.reset:
            cached = checkpoint[key_cp]
            for col in cols_result:
                df.at[idx, col] = cached.get(col)
            continue

        # Construire la requête
        query = build_query(row)
        lat = row.get('Latitude')
        lon = row.get('Longitude')

        print(f"[{idx+1}/{total}] 🔍 {nom[:45]:<45} | {row.get('Ville','')}")

        # Appel API
        results = search_gmaps(query, lat, lon, api_key)
        time.sleep(args.rate)

        if args.debug and results is not None:
            print(f"         🔎 Réponse brute : {json.dumps(results, ensure_ascii=False)[:500]}")

        if results is None:
            erreurs += 1
            metrics = extract_metrics(None)
        else:
            best = extract_best_match(results, nom, lat, lon)
            metrics = extract_metrics(best)

        # Calculer score taille
        score = score_taille(metrics['gmaps_avis'])
        metrics['score_taille'] = score
        metrics['label_taille'] = label_taille(score)

        # Afficher résultat
        if metrics['gmaps_avis'] is not None:
            print(f"         ✅ {metrics['gmaps_nom'][:40] if metrics['gmaps_nom'] else '?'} "
                  f"| {metrics['gmaps_avis']} avis | ⭐{metrics['gmaps_note']} "
                  f"| {metrics['label_taille']}")
        else:
            print(f"         ⚠️  Aucun résultat trouvé")

        # Mettre à jour le dataframe
        for col in cols_result:
            df.at[idx, col] = metrics.get(col)

        # Sauvegarder dans le checkpoint
        checkpoint[key_cp] = {col: metrics.get(col) for col in cols_result}
        save_checkpoint(checkpoint)
        nouveaux += 1

    # ── RÉSUMÉ ───────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"✅ Enrichissement terminé")
    print(f"   Nouveaux : {nouveaux} | Depuis cache : {deja_faits} | Erreurs : {erreurs}")
    print(f"   Checkpoint sauvegardé : {CHECKPOINT}")

    # Distribution des scores
    print(f"\n📊 Distribution par taille :")
    for s in [5, 4, 3, 2, 1, None]:
        n = (df['score_taille'] == s).sum() if s is not None else df['score_taille'].isna().sum()
        print(f"   {label_taille(s)} : {n}")

    # ── EXPORT CSV ───────────────────────────────────────────────────────────
    df.to_csv(OUTPUT_CSV, index=False, encoding='utf-8-sig')
    print(f"\n💾 CSV exporté : {OUTPUT_CSV}")

    # ── EXPORT EXCEL ENRICHI ─────────────────────────────────────────────────
    print(f"\n📝 Génération de l'Excel enrichi : {OUTPUT_EXCEL}...")
    export_excel(df, OUTPUT_EXCEL)
    print(f"✅ Fichier prêt : {OUTPUT_EXCEL}")

# ── EXPORT EXCEL ─────────────────────────────────────────────────────────────
def export_excel(df, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    DARK_BLUE = '1F3864'; MED_BLUE = '2E75B6'; WHITE = 'FFFFFF'
    YELLOW = 'FFD966'; LIGHT_BLUE = 'DEEAF1'

    def thin():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()
    ws = wb.active
    ws.title = '🎯 Prospects enrichis'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    cols = [
        ('Bannière', 12), ('Groupe / Propriétaire', 28), ('Nom magasin', 35),
        ('Ville', 18), ('Région', 20), ('Score territoire', 10),
        ('Score taille\n(Google)', 10), ('Label taille', 22),
        ('Nb avis\nGoogle', 10), ('Note\nGoogle', 8),
        ('Format', 18), ('Contact groupe', 35),
        ('Téléphone', 16), ('URL magasin', 35), ('URL Google Maps', 35),
        ('Nom Google Maps', 30), ('Adresse Google', 35),
        ('Place ID', 28), ('Notes SDR', 30),
    ]

    # Titre
    ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
    c = ws['A1']
    c.value = 'PROSPECTS INTERBOIS — Enrichi Google Maps'
    c.font = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill = PatternFill('solid', start_color=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # En-têtes
    for ci, (h, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        cell = ws.cell(2, ci)
        cell.value = h
        cell.font = Font(name='Arial', bold=True, size=9, color=WHITE)
        cell.fill = PatternFill('solid', start_color=MED_BLUE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin()
    ws.row_dimensions[2].height = 30

    # Tri : score taille desc, score territoire desc
    df_sorted = df.copy()
    df_sorted['_st'] = pd.to_numeric(df_sorted.get('score_taille'), errors='coerce').fillna(0)
    df_sorted['_sterr'] = pd.to_numeric(df_sorted.get('Score_territoire'), errors='coerce').fillna(0)
    df_sorted = df_sorted.sort_values(['_st', '_sterr'], ascending=[False, False])

    score_colors = {5: '70AD47', 4: 'B4D7A8', 3: 'FFD966', 2: 'F4B942', 1: 'FF9999'}

    for ri, (_, row) in enumerate(df_sorted.iterrows()):
        er = ri + 3
        score_t = row.get('score_taille')
        bg = score_colors.get(score_t, 'FFFFFF') if score_t else 'F2F2F2'

        vals = [
            row.get('Bannière',''), row.get('Groupe',''), row.get('Nom',''),
            row.get('Ville',''), row.get('Région',''), row.get('Score_territoire',''),
            row.get('score_taille',''), row.get('label_taille',''),
            row.get('gmaps_avis',''), row.get('gmaps_note',''),
            row.get('Format',''), row.get('Contact_groupe',''),
            row.get('Téléphone',''), row.get('URL_magasin',''), row.get('URL_GoogleMaps',''),
            row.get('gmaps_nom',''), row.get('gmaps_adresse',''),
            row.get('gmaps_place_id',''), '',
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(er, ci)
            c.value = val if val is not None and str(val) != 'nan' else ''
            c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(vertical='center')
            c.border = thin()
        ws.row_dimensions[er].height = 15

    ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}{len(df_sorted)+2}'

    # Onglet légende
    ws2 = wb.create_sheet('ℹ️ Légende scores')
    ws2['A1'] = 'Score taille (proxy nb avis Google)'
    ws2['A1'].font = Font(bold=True, name='Arial', size=11)
    legend = [
        (5, '🟢 Grand', '500+ avis', 'Grand format, fort volume contracteurs'),
        (4, '🔵 Moyen-grand', '200-499 avis', 'Format intéressant'),
        (3, '🟡 Moyen', '100-199 avis', 'Format standard'),
        (2, '🟠 Petit-moyen', '50-99 avis', 'Vérifier la capacité d\'exposition'),
        (1, '🔴 Petit', '<50 avis', 'Probablement hors cible'),
    ]
    ws2.append(['Score', 'Label', 'Seuil avis', 'Interprétation'])
    for row in legend:
        ws2.append(list(row))
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 40

    wb.save(path)

if __name__ == '__main__':
    main()
